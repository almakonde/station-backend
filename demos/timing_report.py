import requests

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

demo_url = "http://127.0.0.1:5003/examinations"


def excel_duration_time(seconds: float):
    return seconds/(3600*24)
    
def key_timings(ex_timing: dict):
    kt = {}
    if 'ongoing' in ex_timing.keys() and 'done' in ex_timing.keys():
        kt['cycle'] = ex_timing['done'] - ex_timing['ongoing'] 
    
    if 'ongoing' in ex_timing.keys() and 'waiting_patient' in ex_timing.keys():
        kt['patient_wait'] = ex_timing['ongoing'] - ex_timing['waiting_patient']

    if 'preparing' in ex_timing.keys() and 'waiting_patient' in ex_timing.keys():
        kt['preparation'] = ex_timing['waiting_patient'] - ex_timing['preparing']
    
    if 'preparing' in ex_timing.keys() and 'pool' in ex_timing.keys():
        kt['fetch'] = ex_timing['preparing'] - ex_timing['pool']
    
    if 'done' in ex_timing.keys() and 'pool' in ex_timing.keys():
        kt['total'] = ex_timing['done'] - ex_timing['pool']
    
    return kt

def generate(url: str, report_filepath):
    resp = requests.get(url)
    exs = resp.json()
    kts = {}

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'timings'
    
    for exid, ex in exs.items():
        if 'timing' in ex.keys():
            kt = key_timings(ex['timing'])
            kts[exid] = kt

    kts_headers = { key for exid, kt in kts.items() for key in kt.keys()}
    
    ex_headers = ['Examination ID', 'Patient ID', 'Name'   ]

    ex_headers_to_keys = {'Examination ID':'examination_id', 'Patient ID':'patient_id', 'Name':'patient_name'}

    headers = [*ex_headers, *kts_headers]

    headers_col = {}

    for i, header in enumerate(headers, start=1):
        ws1.cell(1, i).value = header
        headers_col[header] = i

    row = 2
    for exid, ex in exs.items():
        for header in headers:
            col = headers_col[header]
            if header in ex_headers:
                ws1.cell(row, col).value = ex[ex_headers_to_keys[header]]
            if header in kts_headers:
                if exid in kts.keys():
                    if header in kts[exid].keys():
                        ws1.cell(row, col).value = excel_duration_time(kts[exid][header])
                        ws1.cell(row, col).number_format = '[mm]:ss'
            
        row += 1
    wb.save(report_filepath)



generate(demo_url, './report.xlsx')
